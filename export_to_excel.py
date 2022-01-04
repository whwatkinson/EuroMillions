from datetime import datetime
from enum import Enum

from openpyxl import Workbook

from draw import Draw
from lucky_dip_ticket import LuckDipTicketList, ExportTicket
from tuesday import Tuesday

TICKETS = 1000


class Sheets(Enum):
    DRAW_RESULTS = "DRAW_RESULTS"
    TICKET_RESULTS_ALL = "TICKET_RESULTS_ALL"
    TICKET_RESULTS_WINNERS = "TICKET_RESULTS_WINNERS"
    METADATA = "METADATA"


def setup_csv_file() -> Workbook:
    # TODO make this dynamic
    wb = Workbook()

    fields = ExportTicket.__fields__.keys()

    # Draw results
    sheet = Sheets.DRAW_RESULTS.value
    wb.active.title = sheet
    data = wb[sheet]
    active_row = 1

    data.cell(row=active_row, column=1).value = "draw_date"
    data.cell(row=active_row, column=2).value = "uuid"
    data.cell(row=active_row, column=3).value = "total_prize_money"
    data.cell(row=active_row, column=4).value = "main_numbers"
    data.cell(row=active_row, column=5).value = "lucky_numbers"

    # Ticket results
    sheet = Sheets.TICKET_RESULTS_ALL.value
    wb.create_sheet(sheet)
    data = wb[sheet]
    active_row = 1
    # TODO make this dynamic
    data.cell(row=active_row, column=1).value = "uuid"
    data.cell(row=active_row, column=2).value = "ticket_cost"
    data.cell(row=active_row, column=3).value = "main_numbers"
    data.cell(row=active_row, column=4).value = "main_matches_count"
    data.cell(row=active_row, column=5).value = "main_matches"
    data.cell(row=active_row, column=6).value = "lucky_numbers"
    data.cell(row=active_row, column=7).value = "lucky_matches_count"
    data.cell(row=active_row, column=8).value = "lucky_matches"
    data.cell(row=active_row, column=9).value = "winner"
    data.cell(row=active_row, column=10).value = "has_all_main_numbers"
    data.cell(row=active_row, column=11).value = "has_both_lucky_numbers"
    data.cell(row=active_row, column=12).value = "total_matches"
    data.cell(row=active_row, column=13).value = "prize"
    data.cell(row=active_row, column=14).value = "prize_identifier"

    # Ticket results winners
    sheet = Sheets.TICKET_RESULTS_WINNERS.value
    wb.create_sheet(sheet)
    data = wb[sheet]
    active_row = 1

    data.cell(row=active_row, column=1).value = "uuid"
    data.cell(row=active_row, column=2).value = "ticket_cost"
    data.cell(row=active_row, column=3).value = "main_numbers"
    data.cell(row=active_row, column=4).value = "main_matches_count"
    data.cell(row=active_row, column=5).value = "main_matches"
    data.cell(row=active_row, column=6).value = "lucky_numbers"
    data.cell(row=active_row, column=7).value = "lucky_matches_count"
    data.cell(row=active_row, column=8).value = "lucky_matches"
    data.cell(row=active_row, column=9).value = "winner"
    data.cell(row=active_row, column=10).value = "has_all_main_numbers"
    data.cell(row=active_row, column=11).value = "has_both_lucky_numbers"
    data.cell(row=active_row, column=12).value = "total_matches"
    data.cell(row=active_row, column=13).value = "prize"
    data.cell(row=active_row, column=14).value = "prize_identifier"

    return wb


def add_results_csv(workbook: Workbook, event: Tuesday):

    draw = event.draw
    tickets = event.tickets

    data = workbook[Sheets.DRAW_RESULTS.value]
    active_row = 2
    data.cell(row=active_row, column=1).value = draw.draw_date
    data.cell(row=active_row, column=2).value = str(draw.uuid)
    data.cell(row=active_row, column=3).value = str(draw.total_prize_money)
    data.cell(row=active_row, column=4).value = str(draw.main_numbers)
    data.cell(row=active_row, column=5).value = str(draw.lucky_numbers)

    # all tickets
    data = workbook[Sheets.TICKET_RESULTS_ALL.value]
    active_row = 2
    for ticket in tickets.tickets:
        export = ticket.prepare_ticket_for_export()
        for i, item in enumerate(export, 1):
            k, v = item
            if type(v) is set and len(v) == 0:
                v = ""
            data.cell(row=active_row, column=i).value = str(v)

        active_row += 1

    # winner tickets

    data = workbook[Sheets.TICKET_RESULTS_WINNERS.value]
    active_row = 2
    winners = event.get_winners()

    for ticket in winners:
        export = ticket.prepare_ticket_for_export()
        for i, item in enumerate(export, 1):
            k, v = item
            if type(v) is set and len(v) == 0:
                v = ""
            data.cell(row=active_row, column=i).value = str(v)

        active_row += 1


def data_to_csv(res: Tuesday):

    # create wb
    time_created = datetime.now()
    file_name = f"results {time_created}"
    wb = setup_csv_file()

    add_results_csv(wb, res)

    # Data to wb

    # Save wb
    wb.save(f"/Users/harrywatkinson/Desktop/{file_name}.xlsx")


if __name__ == "__main__":

    ticket_list = LuckDipTicketList(number_of_tickets=TICKETS)
    the_draw = Draw()
    the_draw.auto_draw_all()
    w = Tuesday(the_draw, ticket_list)
    w.check_results()

    data_to_csv(w)
